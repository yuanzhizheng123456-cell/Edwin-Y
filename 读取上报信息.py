from datetime import datetime, timedelta
import binascii
import os
import sys
import threading
import serial
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, simpledialog, ttk

import serial.tools.list_ports
import time


# ports = list(serial.tools.list_ports.comports())
# for p in ports: p.device
# serial = serial.Serial(p.device,9600,timeout=1)

ser = None
stop_event = threading.Event()


def check_stop_requested():
    if stop_event.is_set():
        raise RuntimeError('用户强制停止')

work = "1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F1F"
time_read = 'A5FFFFFFFFFFFFFF99999900068112000000073BBB5A'

report_time = 50
signal_threshold_strong = -77
signal_threshold_good = -89
signal_threshold_normal = -99
param_unlock_password = 'yuan'
# keep_alive_interval = 5




receive_data = ''
frame_header = ''   # ABNT
report_list = []
report_list_before = []
report_list_after = []
report_time_points = ['/' for _ in range(16)]
report_list_num = ['获取CCID失败','附着网络失败','连接服务器失败','发送数据失败','获取IMEI失败','NBAND设置失败','未定义类型','上报成功次数',
                   '主动上报','手动上报','失败重试','关阀24h上报','即时告警','周期数据','非上述类上报','上报总数' ]
report_code_list = [f'ERR{i}' for i in range(8)] + [f'NUM{i}' for i in range(8)]
current_meter_no = 'UNKNOWN'
summary_info = {
    'report_result': '未执行',
    'final_message': '',
    'valve_required': False,
    'valve_result': '未执行',
    'close_valve_result': '未执行',
    'open_valve_result': '未执行',
    'abnt': '',
    'rsrp': '',
    'snr': '',
    'signal_level': '未评估',
    'signal_color': '#e0e0e0',
    'success_count': '',
    'imei': '',
    'ccid': ''
}


def reset_summary_info():
    summary_info.update({
        'report_result': '未执行',
        'final_message': '',
        'valve_required': False,
        'valve_result': '未执行',
        'close_valve_result': '未执行',
        'open_valve_result': '未执行',
        'abnt': '',
        'rsrp': '',
        'snr': '',
        'signal_level': '未评估',
        'signal_color': '#e0e0e0',
        'success_count': '',
        'imei': '',
        'ccid': ''
    })


def classify_signal_strength(rsrp_value):
    try:
        rsrp = int(rsrp_value)
    except Exception:
        return '未知', '#bdbdbd'

    if rsrp > signal_threshold_strong:
        return '特别强', '#6fcf97'
    if rsrp >= signal_threshold_good:
        return '强', '#8bc34a'
    if rsrp >= signal_threshold_normal:
        return '一般', '#ffd54f'
    return '弱', '#ff8a65'


def tag_22_parsing(receive_data):
    data_point = 6
    data_len = int(receive_data[2:6], 16) if len(receive_data) >= 6 else 0
    esim_len = 0

    while data_point < min(len(receive_data), data_len * 2 + 6):
        tag = receive_data[data_point:data_point + 2].upper()
        if len(tag) < 2:
            break

        if tag == '11' and data_point + 32 <= len(receive_data):
            imei = bytes.fromhex(receive_data[data_point + 2:data_point + 32]).decode('ascii', errors='ignore')
            summary_info['imei'] = imei
            print('IMEI号: ', imei)
            data_point += 32
        elif tag == '12' and data_point + 4 <= len(receive_data):
            esim_len = int(receive_data[data_point + 2:data_point + 4], 16)
            data_point += 4
        elif tag == '13':
            end_point = data_point + 2 + esim_len * 2
            if esim_len > 0 and end_point <= len(receive_data):
                ccid = bytes.fromhex(receive_data[data_point + 2:end_point]).decode('ascii', errors='ignore')
                summary_info['ccid'] = ccid
                print('CCID: ', ccid)
                data_point = end_point
            else:
                data_point += 2
        else:
            # 未使用详细解析时，向后跳过一个字节避免死循环。
            data_point += 2


#  水表唤醒
def meter_work(max_retries=5):
    for attempt in range(1, max_retries + 1):
        check_stop_requested()
        ser.write(bytes.fromhex(work))
        time.sleep(2)
        check_stop_requested()
        ser.write(bytes.fromhex(time_read))
        a = binascii.b2a_hex(ser.read(50)).decode()
        if len(a) > 0:
            c = '前导触发成功，开始测试'
            print(c)
            return c

        print(f'等待唤醒({attempt}/{max_retries})')

    raise RuntimeError('无通讯')

def tag_22_parsing(receive_data):
    tag_id = 22
    # print('tag_id: ', tag_id)
    data_point = 6
    data_len = int(receive_data[2:6], 16)
    tag_12 = 0
    while data_len * 2+6 > data_point:
        # 当前总累计流量（正流量-逆流量）
        if receive_data[data_point:data_point+2] == '00':
            tag_00 = int(receive_data[data_point+2:data_point+12], 16)
            if tag_00 > 0x7fffffffff:
                tag_00 -= 0x10000000000
            data_point += 12
            # print('00 当前总累计流量: ', tag_00, 'L')
        # 日结总累计正流量
        elif receive_data[data_point:data_point + 2] == '01':
            tag_01 = int(receive_data[data_point + 2:data_point + 12], 16)
            data_point += 12
            # print('01 日结总累计正流量: ', tag_01, 'L')
        # 日结总累计逆流量
        elif receive_data[data_point:data_point + 2] == '02':
            tag_02 = int(receive_data[data_point + 2:data_point + 12], 16)
            data_point += 12
            # print('02 日结总累计逆流量: ', tag_02, 'L')
        # 采集时间
        elif receive_data[data_point:data_point + 2] == '03':
            tag_03 = receive_data[data_point + 2:data_point + 14]
            data_point += 14
            # print('03 采集时间: ', tag_03)
        # 日最高流量
        elif receive_data[data_point:data_point + 2] == '04':
            tag_04 = int(receive_data[data_point + 2:data_point + 6], 16)
            data_point += 6
            # print('04 日最高流量: ', tag_04, 'L/h')
        # 日最高流量时间
        elif receive_data[data_point:data_point + 2] == '05':
            tag_05 = receive_data[data_point + 2:data_point + 14]
            data_point += 14
            # print('05 日最高流量时间: ', tag_05)
        # 电池电压
        elif receive_data[data_point:data_point + 2] == '06':
            tag_06 = int(receive_data[data_point + 2:data_point + 6], 16)/100
            data_point += 6
            # print('06 电池电压: ', tag_06, 'V')
        # 水表状态
        elif receive_data[data_point:data_point + 2] == '09':
            tag_09 = receive_data[data_point + 2:data_point + 10]
            # print('09 水表状态字: ', tag_09)
            # meter_status_parsing(tag_09)
            data_point += 10
        # 管内水压
        elif receive_data[data_point:data_point + 2] == '0A' or receive_data[data_point:data_point + 2] == '0a':
            if receive_data[data_point + 2:data_point + 6] == 'FFFF':
                tag_0A = '无法获取管内水压'
                # print('无法获取管内水压')
            else:
                tag_0A = int(receive_data[data_point + 2:data_point + 6], 16)/100
                # print('0A 管内水压: ', tag_0A)
            data_point += 6
        # 剩余水量
        elif receive_data[data_point:data_point + 2] == '0B' or receive_data[data_point:data_point + 2] == '0b':
            tag_0B = int(receive_data[data_point + 2:data_point + 12], 16)/1000
            data_point += 12
            # print('0B 剩余水量: ', tag_0B, 'm³')
        # 账户余额
        elif receive_data[data_point:data_point + 2] == '0C' or receive_data[data_point:data_point + 2] == '0c':
            tag_0C = int(receive_data[data_point + 2:data_point + 18], 16)
            data_point += 18
            # print('0C 账户余额: ', tag_0C)
        # 当前总累积正流量
        elif receive_data[data_point:data_point + 2] == '0D' or receive_data[data_point:data_point + 2] == '0d':
            tag_0D = int(receive_data[data_point + 2:data_point + 12], 16)
            data_point += 12
            # print('0D 当前总累积正流量: ', tag_0D, 'L')
        # 当前总累积逆流量
        elif receive_data[data_point:data_point + 2] == '0E' or receive_data[data_point:data_point + 2] == '0e':
            tag_0E = int(receive_data[data_point + 2:data_point + 12], 16)
            data_point += 12
            # print('0E 当前总累积逆流量: ', tag_0E, 'L')
        # 日冻结时间
        elif receive_data[data_point:data_point + 2] == '0F' or receive_data[data_point:data_point + 2] == '0f':
            tag_0F = receive_data[data_point + 2:data_point + 14]
            data_point += 14
            # print('0F 日冻结时间: ', tag_0F)
        # 日结剩余水量
        elif receive_data[data_point:data_point + 2] == '10':
            tag_10 = int(receive_data[data_point + 2:data_point + 12], 16)
            if tag_10 > 0x7fffffffff:
                tag_10 -= 0x10000000000
            data_point += 12
            # print('10 日结剩余水量: ', tag_10, 'L')
        elif receive_data[data_point:data_point + 2] == '11':
            tag_11 = bytes.fromhex(receive_data[data_point + 2:data_point + 32]).decode('ascii', errors='ignore')
            data_point += 32
            summary_info['imei'] = tag_11
            print('IMEI号: ', tag_11)
        elif receive_data[data_point:data_point + 2] == '12':
            tag_12 = int(receive_data[data_point + 2:data_point + 4], 16)
            data_point += 4
            # print('12 eSIM长度: ', tag_12)
        elif receive_data[data_point:data_point + 2] == '13':
            tag_13 = bytes.fromhex(receive_data[data_point + 2:data_point + tag_12*2+2]).decode('ascii', errors='ignore')
            data_point += tag_12*2+2
            summary_info['ccid'] = tag_13
            print('CCID: ', tag_13)
        elif receive_data[data_point:data_point + 2] == '14':
            tag_14 = receive_data[data_point + 2:data_point + 14]
            data_point += 14
            # print('14 最近一次曲线捕获时间: ', tag_14)
        elif receive_data[data_point:data_point + 2] == '15':
            tag_15 = int(receive_data[data_point + 2:data_point + 4]+receive_data[data_point + 4:data_point + 6], 16)
            data_point += 6
            # print('15 曲线数据长度: ', tag_15)
        elif receive_data[data_point:data_point + 2] == '16':
            if tag_15 == 0:
                data_point += 2
            else:
                tag_16 = receive_data[data_point + 2:data_point + tag_15*2+2]
                tag_16_1 = int(tag_16[0:4], 16)  # 捕获周期
                tag_16_2 = int(tag_16[4:6], 16)  # 点数
                tag_16_3 = int(tag_16[6:14], 16)  # 首次捕获时间
                base_time = datetime(1970, 1, 1, 0, 0, 0)
                target_time = base_time + timedelta(seconds=tag_16_3)
                tag_16_4 = int(tag_16[14:24], 16)  # 首次总累计用水量
                if tag_16_4 >= 0x8000000000:
                    tag_16_4 -= 0x10000000000
                # print('16 捕获周期: ', tag_16_1, 'min' ,'点数：', tag_16_2, '首次捕获时间戳：', tag_16_3, '转换后时间：', target_time, '首次总累计用水量：', tag_16_4 ,'L')
                end_point = (tag_15*2 - 24)//4
                for i in range(end_point):
                    increment = int(tag_16[24 + i * 4:24 + i * 4 + 4], 16)
                    if increment > 0x7fff:
                        increment = increment - 0x10000
                    # print('第 ', i+2, ' 次用水累积增量：', increment, 'L')
                data_point += tag_15*2+2
        elif receive_data[data_point:data_point + 2] == '17':
            tag_17 = int(receive_data[data_point + 4:data_point + 6], 16)
            data_point += 6
            # print('17 日最低温度: ', tag_17)
        elif receive_data[data_point:data_point + 2] == '18':
            tag_18 = int(receive_data[data_point + 4:data_point + 6], 16)
            data_point += 6
            # print('18 日最高温度: ', tag_18)
        elif receive_data[data_point:data_point + 2] == '19':
            tag_19 = int(receive_data[data_point + 4:data_point + 6]+receive_data[data_point + 2:data_point + 4], 16)/1000
            data_point += 6
            # print('19 日最低压力: ', tag_19)
        elif receive_data[data_point:data_point + 2] == '1A' or receive_data[data_point:data_point + 2] == '1a':
            tag_1a = int(receive_data[data_point + 4:data_point + 6]+receive_data[data_point + 2:data_point + 4], 16)/1000
            data_point += 6
            # print('1A 日最高压力: ', tag_1a)
        elif receive_data[data_point:data_point + 2] == '1B' or receive_data[data_point:data_point + 2] == '1b':
            # if receive_data[data_point + 2:data_point + 4] == '00':
            #     print('1B NBIOT通讯状态: 在线')
            # elif receive_data[data_point + 2:data_point + 4] == '01':
            #     print('1B NBIOT通讯状态: 离线')
            # elif receive_data[data_point + 2:data_point + 4] == '02':
            #     print('1B NBIOT通讯状态: 故障')
            data_point += 4
        elif receive_data[data_point:data_point + 2] == '1C' or receive_data[data_point:data_point + 2] == '1c':
            # if receive_data[data_point + 2:data_point + 4] == '00':
            #     print('1C 上一次NBIOT通讯状态: 在线')
            # elif receive_data[data_point + 2:data_point + 4] == '01':
            #     print('1C 上一次NBIOT通讯状态: 离线')
            # elif receive_data[data_point + 2:data_point + 4] == '02':
            #     print('1C上一次NBIOT通讯状态: 故障')
            data_point += 4
        elif receive_data[data_point:data_point + 2] == '1D' or receive_data[data_point:data_point + 2] == '1d':
            tag_1d = int(receive_data[data_point + 2:data_point + 4], 16)
            data_point += 4
            # print('1D 电池电量: ', tag_1d, '%')
        elif receive_data[data_point:data_point + 2] == '1E' or receive_data[data_point:data_point + 2] == '1e':
            tag_1e = int(receive_data[data_point + 2:data_point + 4], 16)
            data_point += 4
            # print('1E 固件版本长度: ', tag_1e)
        elif receive_data[data_point:data_point + 2] == '1F' or receive_data[data_point:data_point + 2] == '1f':
            tag_1f = receive_data[data_point + 2:data_point + tag_1e*2+2]
            data_point += tag_1e*2+2
            # print('1F 固件版本号: ', tag_1f)
        elif receive_data[data_point:data_point + 2] == '20':
            tag_20 = receive_data[data_point + 2:data_point + 14]
            data_point += 14
            # print('20 最近OTA更新时间: ', tag_20)
        elif receive_data[data_point:data_point + 2] == '21':
            # if receive_data[data_point + 2:data_point + 4] == '00':
            #     print('21 阀门当前状态：0%开')
            # elif receive_data[data_point + 2:data_point + 4] == '01':
            #     print('21 阀门当前状态：25%开')
            # elif receive_data[data_point + 2:data_point + 4] == '02':
            #     print('21 阀门当前状态：50%开')
            # elif receive_data[data_point + 2:data_point + 4] == '03':
            #     print('21 阀门当前状态：75%开')
            # elif receive_data[data_point + 2:data_point + 4] == '04':
            #     print('21 阀门当前状态：100%开')
            data_point += 4
    print()

# tag22前置解析
def platform_protocol_analysis(frame_data):
    # 水表id号
    receive_data = frame_data

    meter_num = bytes.fromhex(receive_data[2:32]).replace(b'\x00', b'').decode('ascii')
    summary_info['abnt'] = meter_num
    print('ABNT：', meter_num)

    if receive_data[38] == '8':
        meter_rsrp = '-'+receive_data[39:42]
    else:
        meter_rsrp = '+'+receive_data[39:42]
    summary_info['rsrp'] = meter_rsrp
    print('信号强度RSRP:', meter_rsrp)

    signal_level, signal_color = classify_signal_strength(meter_rsrp)
    summary_info['signal_level'] = signal_level
    summary_info['signal_color'] = signal_color

    if receive_data[42] == '8':
        meter_snr = '-'+receive_data[43:46]
    else:
        meter_snr = '+'+receive_data[43:46]
    summary_info['snr'] = meter_snr
    print('信噪比SNR:', meter_snr)


    tlv_len = int(receive_data[64:66], 16)
    tlv_data = receive_data[66:tlv_len*2+72]
    tag_22_parsing(tlv_data)
# 计算累加和
def Cal_sum(string):   #计算校验和

    hex_str = string + 'FFFFFFFF'
    total = 0
    for i in range(0, len(hex_str), 2):
        byte = int(hex_str[i:i+2], 16)
        total += byte
    checksum = total & 0xFF
    return hex_str + format(checksum, '02X') + '16'
# 获取系统时间
def get_time():
    a = str(datetime.now())
    timestamp = a[2:4] + a[5:7] + a[8:10] + a[11:13] + a[14:16] + a[17:19]
    return timestamp
# 发送tag11查询tag22
def send_tag11():
    command_list_header = '0B'
    command_list_data = '1401'

    command_list_len = format(len(command_list_data) // 2, '04X')
    data_list = command_list_header + command_list_len + command_list_data
    data_len1 = format(len(data_list) // 2 + 2, '04X')
    data_len2 = format(len(data_list) // 2, '04X')

    transfer_data = Cal_sum('68' + frame_header + '21' + get_time() + '03' + '00' + data_len1 + data_len2 + data_list )

    ser.write(bytes.fromhex(transfer_data))
    receive_data = binascii.b2a_hex(ser.read(500)).decode()
    # print('tag22: ', receive_data)
    return receive_data
# 获取表号、ABNT
def get_meter_number():
    global frame_header, current_meter_no
    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF9999990006810100000007BE785A')) #读取表号
    meter_num = binascii.b2a_hex(ser.read(500)).decode()
    if meter_num != '' and len(meter_num) == 54:
        current_meter_no = meter_num[34:48]
        print('表号：', current_meter_no)
        # print('获取表号成功')
    else:
        print('读取表号失败')
        get_meter_number()

    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF999999000681800000000CC3A15A')) #读取ABNT号
    abnt_num = binascii.b2a_hex(ser.read(500)).decode()
    if abnt_num != '' and len(abnt_num) == 64:
        print('ABNT：', bytes.fromhex(abnt_num[34:58]).decode('ascii'))
        # print('获取ABNT号成功')
        frame_header = '000000' + abnt_num[34:58]
    else:
        print('读取ABNT失败')
        get_meter_number()
# 获取上报事件数据
def read_report_status(err_status = 0):
    check_stop_requested()
    global report_list, report_time_points
    report_list.clear()
    report_time_points = ['/' for _ in range(16)]
    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF999999000684F00000005082065A'))  # 读取上报事件
    p = binascii.b2a_hex(ser.read(500)).decode()
    if p != '' and len(p) == 200:
        p = p
    else:
        print('读取上报事件失败')
        read_report_status(1)
    temp = 46
    for i in range(8):
        report_list.append(int(p[temp:temp+2], 16))
        temp += 16
    temp = 162
    for i in range(8):
        report_list.append(int(p[temp:temp+2], 16))
        temp += 4

    for i in range(8):
        ts = p[34 + i * 16:46 + i * 16]
        report_time_points[i] = format_report_timestamp(ts)

    if err_status == 1:

        print(
            f'ERR0\t获取CCID失败\t{p[34:36]}-{p[36:38]}-{p[38:40]} {p[40:42]}:{p[42:44]}:{p[44:46]}\t{report_list[0]}\t\tNUM0\t主动上报\t\t{report_list[8]}')
        print(
            f'ERR1\t附着网络失败\t{p[50:52]}-{p[52:54]}-{p[54:56]} {p[56:58]}:{p[58:60]}:{p[60:62]}\t{report_list[1]}\t\tNUM1\t手动上报\t\t{report_list[9]}')
        print(
            f'ERR2\t连接服务器失败\t{p[66:68]}-{p[68:70]}-{p[70:72]} {p[72:74]}:{p[74:76]}:{p[76:78]}\t{report_list[2]}\t\tNUM2\t失败重试\t\t{report_list[10]}')
        print(
            f'ERR3\t发送数据失败\t{p[82:84]}-{p[84:86]}-{p[86:88]} {p[88:90]}:{p[90:92]}:{p[92:94]}\t{report_list[3]}\t\tNUM3\t关阀24h上报\t{report_list[11]}')
        print(
            f'ERR4\t获取IMEI失败\t{p[98:100]}-{p[100:102]}-{p[102:104]} {p[104:106]}:{p[106:108]}:{p[108:110]}\t{report_list[4]}\t\tNUM4\t即时告警\t\t{report_list[12]}')
        print(
            f'ERR5\tNBAND设置失败\t{p[114:116]}-{p[116:118]}-{p[118:120]} {p[120:122]}:{p[122:124]}:{p[124:126]}\t{report_list[5]}\t\tNUM5\t周期数据\t\t{report_list[13]}')
        print(
            f'ERR6\t未定义类型\t{p[130:132]}-{p[132:134]}-{p[134:136]} {p[136:138]}:{p[138:140]}:{p[140:142]}\t{report_list[6]}\t\tNUM6\t非上述类上报\t{report_list[14]}')
        print(
            f'ERR7\t上报成功次数\t{p[146:148]}-{p[148:150]}-{p[150:152]} {p[152:154]}:{p[154:156]}:{p[156:158]}\t{report_list[7]}\t\tNUM7\t上报总数\t\t{report_list[15]}\t ')


def format_report_timestamp(ts_hex):
    if len(ts_hex) != 12:
        return '/'
    return f"{ts_hex[0:2]}-{ts_hex[2:4]}-{ts_hex[4:6]} {ts_hex[6:8]}:{ts_hex[8:10]}:{ts_hex[10:12]}"


def build_report_snapshot():
    return [(report_time_points[i], report_list[i]) for i in range(16)]
# 触发升级
def triger_report(on_wait_progress=None):
    check_stop_requested()
    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF999999000506660001668D5D5A'))  # 触发上报
    status = binascii.b2a_hex(ser.read(500)).decode()
    if status != '' and status[30:32] == '01':
        print('触发上报成功')
        print('等待上报完成时间：', report_time, 's')
        for remaining in range(report_time, 0, -1):
            check_stop_requested()
            if on_wait_progress:
                on_wait_progress('上报等待', remaining, report_time, '')
            time.sleep(1)
        if on_wait_progress:
            on_wait_progress('上报等待', 0, report_time, '')
        meter_work()
    else:
        print('触发上报失败')
        triger_report(on_wait_progress)
# 上报事件数据对比
def report_list_compare(max_retries=3, wait_seconds=15, on_wait_progress=None, on_after_snapshot=None):
    global report_list_before, report_list_after
    retry_count = 0
    while report_list_before == report_list_after and retry_count < max_retries:
        check_stop_requested()
        retry_count += 1
        print(f'上报事件未更新，等待更新上报事件({retry_count}/{max_retries})')
        for remaining in range(wait_seconds, 0, -1):
            check_stop_requested()
            if on_wait_progress:
                on_wait_progress('事件更新等待', remaining, wait_seconds, f'{retry_count}/{max_retries}')
            time.sleep(1)
        if on_wait_progress:
            on_wait_progress('事件更新等待', 0, wait_seconds, f'{retry_count}/{max_retries}')
        read_report_status(1)  # 再次获取当前上报事件
        report_list_after = report_list.copy()
        if on_after_snapshot:
            on_after_snapshot(build_report_snapshot())

    if report_list_before == report_list_after:
        print('上报时间长，失败')
        summary_info['report_result'] = '上报失败'
        summary_info['final_message'] = '上报时间长，失败'
        raise RuntimeError('上报时间长，失败')

    for i in range(16):
        if report_list_before[i] != report_list_after[i]:
            print(report_list_num[i], '次数+' + str(report_list_after[i] - report_list_before[i]))
    if report_list_after[7] - report_list_before[7] == report_list_after[15] - report_list_before[15]:
        print('上报成功')
        summary_info['report_result'] = '上报成功'
        summary_info['final_message'] = '上报事件对比通过，上报成功。'
        return True

    print('上报失败')
    summary_info['report_result'] = '上报失败'
    summary_info['final_message'] = '上报事件对比不一致，上报失败。'
    return False


def valve_test(wait_seconds=15, on_wait_progress=None):
    check_stop_requested()
    summary_info['close_valve_result'] = '未执行'
    summary_info['open_valve_result'] = '未执行'

    def wait_with_progress(phase_name):
        print(f'等待{phase_name}：{wait_seconds}s')
        for remaining in range(wait_seconds, 0, -1):
            check_stop_requested()
            if on_wait_progress:
                on_wait_progress(phase_name, remaining, wait_seconds, '')
            time.sleep(1)
        if on_wait_progress:
            on_wait_progress(phase_name, 0, wait_seconds, '')

    # 关阀
    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF999999000502280001A52BE45A'))
    close_ack = binascii.b2a_hex(ser.read(500)).decode()
    close_cmd_ok = bool(close_ack) and len(close_ack) >= 32 and close_ack[30:32] == '01'
    if not close_cmd_ok:
        print('关阀失败')
        summary_info['close_valve_result'] = '失败'
        return False

    print('关阀成功')
    wait_with_progress('关阀等待')
    meter_work()

    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF999999000684B40000000473F65A'))
    close_state = binascii.b2a_hex(ser.read(500)).decode()
    close_state_ok = bool(close_state) and len(close_state) >= 42 and close_state[34:42] == '22000500'
    if close_state_ok:
        print('关阀状态读取成功')
        summary_info['close_valve_result'] = '成功'
    else:
        print('关阀状态失败')
        summary_info['close_valve_result'] = '失败'

    # 开阀
    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF9999990005022800015A6BA45A'))
    open_ack = binascii.b2a_hex(ser.read(500)).decode()
    open_cmd_ok = bool(open_ack) and len(open_ack) >= 32 and open_ack[30:32] == '01'
    if not open_cmd_ok:
        print('开阀失败')
        summary_info['open_valve_result'] = '失败'
        return False

    print('开阀成功')
    wait_with_progress('开阀等待')
    meter_work()

    ser.write(bytes.fromhex('A5FFFFFFFFFFFFFF999999000684B40000000473F65A'))
    open_state = binascii.b2a_hex(ser.read(500)).decode()
    open_state_ok = bool(open_state) and len(open_state) >= 42 and open_state[34:42] == '00000500'
    if open_state_ok:
        print('开阀状态读取成功')
        summary_info['open_valve_result'] = '成功'
    else:
        print('开阀状态失败')
        summary_info['open_valve_result'] = '失败'

    return close_state_ok and open_state_ok


def list_available_ports():
    return [p.device for p in serial.tools.list_ports.comports()]


def open_serial_port(port_name, baud_rate=9600):
    global ser
    if ser and ser.is_open:
        ser.close()
    ser = serial.Serial(port_name, int(baud_rate), timeout=1)


def close_serial_port():
    global ser
    if ser and ser.is_open:
        ser.close()


def run_report_flow(on_before_snapshot=None, on_after_snapshot=None, on_summary_ready=None, on_wait_progress=None, enable_valve_test=False):
    global report_list_before, report_list_after
    check_stop_requested()
    reset_summary_info()
    summary_info['valve_required'] = bool(enable_valve_test)
    meter_work()
    check_stop_requested()
    get_meter_number()                        # 获取表号，ABNT号
    check_stop_requested()
    read_report_status(1)                     # 读取当前上报事件
    report_list_before = report_list.copy()
    before_snapshot = build_report_snapshot()
    if on_before_snapshot:
        on_before_snapshot(before_snapshot)

    triger_report(on_wait_progress=on_wait_progress)  # 触发上报并等待设定时间
    check_stop_requested()
    read_report_status(1)                     # 上报完成再次读取当前上报事件
    report_list_after = report_list.copy()
    if len(report_list_after) > 7:
        summary_info['success_count'] = str(report_list_after[7])
    after_snapshot = build_report_snapshot()
    if on_after_snapshot:
        on_after_snapshot(after_snapshot)

    try:
        report_ok = report_list_compare(on_wait_progress=on_wait_progress, on_after_snapshot=on_after_snapshot)  # 对比上报事件信息
    except RuntimeError as ex:
        # 勾选阀门测试时，允许在上报失败后继续执行阀门测试，再做联合判定。
        if enable_valve_test:
            print(str(ex))
            report_ok = False
            summary_info['report_result'] = '上报失败'
            summary_info['final_message'] = str(ex)
        else:
            raise
    if on_after_snapshot:
        on_after_snapshot(build_report_snapshot())
    platform_protocol_analysis(send_tag11())  # tag11查询tag22数据

    # 上报完成后先刷新中间统计框，阀门测试可继续在后台执行。
    if on_summary_ready:
        on_summary_ready(dict(summary_info))

    if enable_valve_test:
        print('已勾选阀门测试，开始执行阀门测试...')
        valve_ok = valve_test(on_wait_progress=on_wait_progress)
        summary_info['valve_result'] = '成功' if valve_ok else '失败'

        if report_ok and valve_ok:
            summary_info['report_result'] = '测试成功'
            summary_info['final_message'] = '上报成功且阀门测试成功。'
        elif report_ok and (not valve_ok):
            summary_info['report_result'] = '测试失败'
            summary_info['final_message'] = '上报成功，但阀门测试失败。'
        elif (not report_ok) and valve_ok:
            summary_info['report_result'] = '测试失败'
            summary_info['final_message'] = '上报失败，阀门测试成功。'
        else:
            summary_info['report_result'] = '测试失败'
            summary_info['final_message'] = '上报失败且阀门测试失败。'
    else:
        summary_info['valve_result'] = '未执行'
        summary_info['close_valve_result'] = '未执行'
        summary_info['open_valve_result'] = '未执行'

    if enable_valve_test and on_summary_ready:
        on_summary_ready(dict(summary_info))
    return before_snapshot, after_snapshot


class TextRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.line_start = True

    def write(self, message):
        if not message:
            return
        self.text_widget.after(0, self._append, message)

    def flush(self):
        return

    def _append(self, message):
        formatted = self._with_timestamp(message)
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, formatted)
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')

    def _with_timestamp(self, message):
        output = []
        for ch in message:
            if self.line_start and ch != '\n':
                output.append(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ")
                self.line_start = False
            output.append(ch)
            if ch == '\n':
                self.line_start = True
        return ''.join(output)


class UpperComputerApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Report Event Test Tool')
        self.root.geometry('1240x840')
        try:
            self.root.state('zoomed')
        except Exception:
            pass

        self.original_stdout = sys.stdout
        self.original_stderr = sys.stderr
        self.is_running = False
        self.params_unlocked = False

        top_frame = ttk.Frame(root, padding=10)
        top_frame.pack(fill='x')

        body_container = ttk.Frame(root)
        body_container.pack(fill='both', expand=True)

        self.main_canvas = tk.Canvas(body_container, highlightthickness=0)
        self.main_canvas.pack(side='left', fill='both', expand=True)
        body_scrollbar = ttk.Scrollbar(body_container, orient='vertical', command=self.main_canvas.yview)
        body_scrollbar.pack(side='right', fill='y')
        self.main_canvas.configure(yscrollcommand=body_scrollbar.set)

        self.scroll_frame = ttk.Frame(self.main_canvas)
        self.scroll_window_id = self.main_canvas.create_window((0, 0), window=self.scroll_frame, anchor='nw')

        def _on_scroll_frame_configure(event):
            self.main_canvas.configure(scrollregion=self.main_canvas.bbox('all'))

        def _on_canvas_configure(event):
            self.main_canvas.itemconfigure(self.scroll_window_id, width=event.width)

        self.scroll_frame.bind('<Configure>', _on_scroll_frame_configure)
        self.main_canvas.bind('<Configure>', _on_canvas_configure)

        ttk.Label(top_frame, text='串口:').grid(row=0, column=0, padx=4, pady=4, sticky='w')
        self.port_var = tk.StringVar()
        self.port_combo = ttk.Combobox(top_frame, textvariable=self.port_var, width=12, state='readonly')
        self.port_combo.grid(row=0, column=1, padx=4, pady=4, sticky='w')

        ttk.Button(top_frame, text='刷新串口', command=self.refresh_ports).grid(row=0, column=2, padx=4, pady=4)

        ttk.Label(top_frame, text='波特率:').grid(row=0, column=3, padx=4, pady=4, sticky='w')
        self.baud_var = tk.StringVar(value='9600')
        self.baud_combo = ttk.Combobox(
            top_frame,
            textvariable=self.baud_var,
            values=['1200', '2400', '4800', '9600', '19200', '38400', '57600', '115200'],
            width=8,
            state='disabled'
        )
        self.baud_combo.grid(row=0, column=4, padx=4, pady=4, sticky='w')

        ttk.Label(top_frame, text='上报等待时间(s):').grid(row=0, column=5, padx=4, pady=4, sticky='w')
        self.wait_var = tk.StringVar(value=str(report_time))
        self.wait_entry = ttk.Entry(top_frame, textvariable=self.wait_var, width=8, state='disabled')
        self.wait_entry.grid(row=0, column=6, padx=4, pady=4, sticky='w')

        self.unlock_btn = ttk.Button(top_frame, text='解锁参数', command=self.toggle_param_lock)
        self.unlock_btn.grid(row=0, column=7, padx=4, pady=4, sticky='w')

        ttk.Label(top_frame, text='阶梯1(dBm):').grid(row=0, column=8, padx=4, pady=4, sticky='w')
        self.signal_strong_var = tk.StringVar(value=str(signal_threshold_strong))
        self.signal_strong_entry = ttk.Entry(top_frame, textvariable=self.signal_strong_var, width=7, state='disabled')
        self.signal_strong_entry.grid(row=0, column=9, padx=4, pady=4, sticky='w')

        ttk.Label(top_frame, text='阶梯2(dBm):').grid(row=0, column=10, padx=4, pady=4, sticky='w')
        self.signal_good_var = tk.StringVar(value=str(signal_threshold_good))
        self.signal_good_entry = ttk.Entry(top_frame, textvariable=self.signal_good_var, width=7, state='disabled')
        self.signal_good_entry.grid(row=0, column=11, padx=4, pady=4, sticky='w')

        ttk.Label(top_frame, text='阶梯3(dBm):').grid(row=0, column=12, padx=4, pady=4, sticky='w')
        self.signal_normal_var = tk.StringVar(value=str(signal_threshold_normal))
        self.signal_normal_entry = ttk.Entry(top_frame, textvariable=self.signal_normal_var, width=7, state='disabled')
        self.signal_normal_entry.grid(row=0, column=13, padx=4, pady=4, sticky='w')

        ttk.Label(top_frame, text='解锁密码:').grid(row=1, column=8, padx=4, pady=(0, 4), sticky='w')
        self.unlock_pwd_var = tk.StringVar(value=param_unlock_password)
        self.unlock_pwd_entry = ttk.Entry(top_frame, textvariable=self.unlock_pwd_var, width=12, show='*', state='disabled')
        self.unlock_pwd_entry.grid(row=1, column=9, columnspan=2, padx=4, pady=(0, 4), sticky='w')

        self.enable_valve_test_var = tk.BooleanVar(value=False)
        self.valve_check = ttk.Checkbutton(top_frame, text='阀门测试', variable=self.enable_valve_test_var)
        self.valve_check.grid(row=1, column=11, columnspan=3, padx=4, pady=(0, 4), sticky='w')

        self.countdown_var = tk.StringVar(value='等待倒计时: --')
        ttk.Label(top_frame, textvariable=self.countdown_var).grid(row=1, column=0, columnspan=4, padx=4, pady=(0, 4), sticky='w')
        self.wait_phase_var = tk.StringVar(value='等待进度: --')
        ttk.Label(top_frame, textvariable=self.wait_phase_var).grid(row=2, column=0, columnspan=14, padx=4, pady=(0, 4), sticky='w')

        self.start_btn = tk.Button(
            top_frame,
            text='开始测试',
            command=self.start_test,
            font=('Microsoft YaHei', 12, 'bold'),
            bg='#18a558',
            fg='white',
            activebackground='#128247',
            activeforeground='white',
            padx=18,
            pady=6,
            relief='raised',
            bd=2,
            cursor='hand2'
        )
        self.start_btn.grid(row=1, column=4, padx=4, pady=(0, 4), sticky='w')

        self.stop_btn = ttk.Button(top_frame, text='强制停止', command=self.force_stop, state='disabled')
        self.stop_btn.grid(row=1, column=7, padx=4, pady=(0, 4), sticky='w')

        ttk.Button(top_frame, text='清空日志', command=self.clear_log).grid(row=1, column=6, padx=4, pady=(0, 4), sticky='w')

        top_frame.grid_columnconfigure(13, weight=1)
        self.wait_progress = ttk.Progressbar(top_frame, orient='horizontal', mode='determinate', maximum=100)
        self.wait_progress.grid(row=3, column=0, columnspan=14, padx=4, pady=(0, 6), sticky='ew')

        self.status_var = tk.StringVar(value='就绪')
        ttk.Label(self.scroll_frame, textvariable=self.status_var, padding=(10, 0, 10, 8)).pack(anchor='w')

        compare_frame = ttk.Frame(self.scroll_frame)

        self.table = ttk.Treeview(
            compare_frame,
            columns=('code', 'item', 'before_time', 'before_count', 'after_time', 'after_count'),
            show='headings',
            height=10
        )
        self.table.heading('code', text='列表')
        self.table.heading('item', text='项目')
        self.table.heading('before_time', text='上报前-发生时间')
        self.table.heading('before_count', text='上报前-次数')
        self.table.heading('after_time', text='上报后-发生时间')
        self.table.heading('after_count', text='上报后-次数')

        self.table.column('code', width=80, anchor='center')
        self.table.column('item', width=180, anchor='center')
        self.table.column('before_time', width=220, anchor='center')
        self.table.column('before_count', width=120, anchor='center')
        self.table.column('after_time', width=220, anchor='center')
        self.table.column('after_count', width=120, anchor='center')

        table_scroll_y = ttk.Scrollbar(compare_frame, orient='vertical', command=self.table.yview)
        table_scroll_x = ttk.Scrollbar(compare_frame, orient='horizontal', command=self.table.xview)
        self.table.configure(yscrollcommand=table_scroll_y.set, xscrollcommand=table_scroll_x.set)

        self.table.pack(side='left', fill='both', expand=True)
        table_scroll_y.pack(side='right', fill='y')
        table_scroll_x.pack(side='bottom', fill='x')

        self._init_table_rows()

        self.summary_frame = tk.LabelFrame(
            self.scroll_frame,
            text='最终结果统计',
            padx=14,
            pady=12,
            font=('Microsoft YaHei', 10, 'bold'),
            bg='#e0e0e0'
        )
        self.summary_frame.configure(height=260)
        self.summary_frame.pack_propagate(False)
        self.summary_frame.pack(fill='x', expand=False, padx=10, pady=(0, 8))
        compare_frame.pack(fill='x', expand=False, padx=10, pady=(0, 8))

        self.summary_result_var = tk.StringVar(value='未执行')
        self.summary_meter_var = tk.StringVar(value='')
        self.summary_abnt_var = tk.StringVar(value='')
        self.summary_rsrp_var = tk.StringVar(value='')
        self.summary_snr_var = tk.StringVar(value='')
        self.summary_imei_var = tk.StringVar(value='')
        self.summary_ccid_var = tk.StringVar(value='')
        self.summary_success_count_var = tk.StringVar(value='')
        self.summary_message_var = tk.StringVar(value='')

        self.summary_labels = []

        self.summary_frame.grid_columnconfigure(0, weight=0)
        self.summary_frame.grid_columnconfigure(1, weight=0)

        def add_summary_label(text, row, col):
            label = tk.Label(
                self.summary_frame,
                text=text,
                font=('Microsoft YaHei', 10),
                bg='#e0e0e0'
            )
            label.grid(row=row, column=col, padx=6, pady=6, sticky='w')
            self.summary_labels.append(label)

        def add_summary_value(var, row, col, width=16):
            label = tk.Label(
                self.summary_frame,
                textvariable=var,
                width=width,
                anchor='w',
                font=('Consolas', 10),
                bg='#e0e0e0'
            )
            label.grid(row=row, column=col, padx=6, pady=6, sticky='w')
            self.summary_labels.append(label)
            return label

        # 8行纵向显示（去掉“最后打印内容”）
        add_summary_label('表号:', 0, 0)
        add_summary_value(self.summary_meter_var, 0, 1, width=28)

        add_summary_label('ABNT:', 1, 0)
        add_summary_value(self.summary_abnt_var, 1, 1, width=28)

        add_summary_label('信号强度:', 2, 0)
        self.rsrp_value_label = tk.Label(self.summary_frame, textvariable=self.summary_rsrp_var, width=12, relief='groove', font=('Consolas', 10), bg='#e0e0e0')
        self.rsrp_value_label.grid(row=2, column=1, padx=6, pady=6, sticky='w')
        self.summary_labels.append(self.rsrp_value_label)

        add_summary_label('信号等级:', 3, 0)
        self.signal_level_label = tk.Label(self.summary_frame, text='未评估', width=12, relief='groove', font=('Microsoft YaHei', 10), bg='#e0e0e0')
        self.signal_level_label.grid(row=3, column=1, padx=6, pady=6, sticky='w')
        self.summary_labels.append(self.signal_level_label)

        add_summary_label('信噪比:', 4, 0)
        add_summary_value(self.summary_snr_var, 4, 1, width=28)

        add_summary_label('IMEI:', 5, 0)
        add_summary_value(self.summary_imei_var, 5, 1, width=28)

        add_summary_label('CCID:', 6, 0)
        add_summary_value(self.summary_ccid_var, 6, 1, width=48)

        add_summary_label('成功次数:', 7, 0)
        add_summary_value(self.summary_success_count_var, 7, 1, width=28)

        add_summary_label('上报结果:', 8, 0)
        self.result_value_label = tk.Label(self.summary_frame, textvariable=self.summary_result_var, width=12, relief='groove', font=('Microsoft YaHei', 10, 'bold'), bg='#e0e0e0')
        self.result_value_label.grid(row=8, column=1, padx=6, pady=6, sticky='w')
        self.summary_labels.append(self.result_value_label)

        self.log_text = scrolledtext.ScrolledText(self.scroll_frame, wrap='word', font=('Consolas', 10), height=9)
        self.log_text.pack(fill='both', expand=True, padx=10, pady=(0, 10))
        self.log_text.configure(state='disabled')

        sys.stdout = TextRedirector(self.log_text)
        sys.stderr = TextRedirector(self.log_text)

        self.refresh_ports()
        self.root.protocol('WM_DELETE_WINDOW', self.on_close)

    def _init_table_rows(self):
        for i in range(16):
            self.table.insert('', tk.END, iid=f'row_{i}', values=(
                report_code_list[i],
                report_list_num[i],
                '/',
                '',
                '/',
                ''
            ))

    def _clear_table_data(self):
        for i in range(16):
            self.table.item(f'row_{i}', values=(
                report_code_list[i],
                report_list_num[i],
                '/',
                '',
                '/',
                ''
            ))

    def _update_table_data(self, before_snapshot, after_snapshot):
        for i in range(16):
            before_time, before_count = before_snapshot[i]
            after_time, after_count = after_snapshot[i]
            self.table.item(f'row_{i}', values=(
                report_code_list[i],
                report_list_num[i],
                before_time,
                before_count,
                after_time,
                after_count
            ))

    def _update_before_snapshot(self, before_snapshot):
        for i in range(16):
            before_time, before_count = before_snapshot[i]
            self.table.item(f'row_{i}', values=(
                report_code_list[i],
                report_list_num[i],
                before_time,
                before_count,
                '/',
                ''
            ))

    def _update_after_snapshot(self, after_snapshot):
        for i in range(16):
            current_values = self.table.item(f'row_{i}', 'values')
            after_time, after_count = after_snapshot[i]
            self.table.item(f'row_{i}', values=(
                current_values[0],
                current_values[1],
                current_values[2],
                current_values[3],
                after_time,
                after_count
            ))

    def _reset_summary_panel(self):
        self._apply_summary_bg('#ffffff')
        self.summary_result_var.set('未执行')
        self.summary_meter_var.set('')
        self.summary_abnt_var.set('')
        self.summary_rsrp_var.set('')
        self.summary_snr_var.set('')
        self.summary_imei_var.set('')
        self.summary_ccid_var.set('')
        self.summary_success_count_var.set('')
        self.summary_message_var.set('')
        self.rsrp_value_label.configure(bg='#ffffff')
        self.signal_level_label.configure(text='未评估', bg='#ffffff')
        self.result_value_label.configure(bg='#ffffff')

    def _apply_summary_bg(self, bg_color):
        self.summary_frame.configure(bg=bg_color)
        for label in self.summary_labels:
            label.configure(bg=bg_color)

    def _update_summary_panel(self, info):
        self.summary_result_var.set(info.get('report_result', '未执行'))
        self.summary_meter_var.set(current_meter_no)
        self.summary_abnt_var.set(info.get('abnt', ''))
        self.summary_rsrp_var.set(info.get('rsrp', ''))
        self.summary_snr_var.set(info.get('snr', ''))
        self.summary_imei_var.set(info.get('imei', ''))
        self.summary_ccid_var.set(info.get('ccid', ''))
        self.summary_success_count_var.set(info.get('success_count', ''))
        self.summary_message_var.set(info.get('final_message', ''))

        result_text = info.get('report_result', '')
        valve_required = bool(info.get('valve_required', False))
        valve_result = info.get('valve_result', '未执行')

        final_pass = False
        if valve_required:
            # 勾选阀门测试时，只有阀测完成且最终结果成功才变绿。
            final_pass = valve_result != '未执行' and result_text == '测试成功'
        else:
            final_pass = result_text == '上报成功'

        panel_color = '#c8e6c9' if final_pass else '#ffffff'
        self._apply_summary_bg(panel_color)
        self.result_value_label.configure(bg=panel_color)

        color = panel_color
        level = info.get('signal_level', '未评估')
        self.rsrp_value_label.configure(bg=color)
        self.signal_level_label.configure(text=level, bg=color)

    def _update_countdown(self, remaining):
        if remaining > 0:
            self.countdown_var.set(f'等待倒计时: {remaining}s')
        else:
            self.countdown_var.set('等待倒计时: 完成')

    def _update_wait_progress(self, phase, remaining, total, retry_text=''):
        self._update_countdown(remaining)
        total = max(int(total), 1)
        remaining = max(int(remaining), 0)
        done = total - remaining
        self.wait_progress.configure(maximum=total)
        self.wait_progress['value'] = done

        retry_suffix = f' ({retry_text})' if retry_text else ''
        if remaining > 0:
            self.wait_phase_var.set(f'{phase}{retry_suffix}: 剩余 {remaining}s / {total}s')
        else:
            self.wait_phase_var.set(f'{phase}{retry_suffix}: 完成')

    def _show_result_popup(self, result_text):
        if '成功' in result_text:
            bg_color = '#c8e6c9'
            title = '测试结果 - 成功'
        else:
            bg_color = '#ff4d4f'
            title = '测试结果 - 失败'

        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.configure(bg=bg_color)
        popup.transient(self.root)
        popup.grab_set()
        popup.attributes('-topmost', True)
        popup.resizable(False, False)

        # 弹窗面积约为主窗口1/4（宽高各1/2），并设置最小值保证可读性。
        self.root.update_idletasks()
        root_x = self.root.winfo_rootx()
        root_y = self.root.winfo_rooty()
        root_w = self.root.winfo_width()
        root_h = self.root.winfo_height()
        popup_width = max(420, root_w // 2)
        popup_height = max(220, root_h // 2)
        pos_x = root_x + max((root_w - popup_width) // 2, 0)
        pos_y = root_y + max((root_h - popup_height) // 2, 0)
        popup.geometry(f'{popup_width}x{popup_height}+{pos_x}+{pos_y}')

        tk.Label(
            popup,
            text=result_text,
            font=('Microsoft YaHei', 14, 'bold'),
            bg=bg_color,
            padx=30,
            pady=28
        ).pack(expand=True)

        tk.Button(
            popup,
            text='确定',
            width=12,
            command=popup.destroy,
            bg='#ffffff'
        ).pack(pady=(0, 16))

    def refresh_ports(self):
        ports = list_available_ports()
        self.port_combo['values'] = ports
        if ports:
            self.port_combo.current(0)
            self.status_var.set('已获取串口列表')
        else:
            self.port_var.set('')
            self.status_var.set('未找到可用串口')

    def toggle_param_lock(self):
        current_password = self.unlock_pwd_var.get().strip() or 'yuan'
        if self.params_unlocked:
            self._set_param_lock(False)
            self.status_var.set('参数已锁定')
            return

        pwd = simpledialog.askstring('参数解锁', '请输入解锁密码:', show='*', parent=self.root)
        if pwd is None:
            return
        if pwd == current_password:
            self._set_param_lock(True)
            self.status_var.set('参数已解锁，可修改波特率、等待时间、信号阶梯和密码')
        else:
            messagebox.showerror('解锁失败', '密码错误')

    def _set_param_lock(self, unlocked):
        self.params_unlocked = unlocked
        if unlocked:
            self.baud_combo.configure(state='readonly')
            self.wait_entry.configure(state='normal')
            self.signal_strong_entry.configure(state='normal')
            self.signal_good_entry.configure(state='normal')
            self.signal_normal_entry.configure(state='normal')
            self.unlock_pwd_entry.configure(state='normal')
            self.unlock_btn.configure(text='锁定参数')
        else:
            self.baud_combo.configure(state='disabled')
            self.wait_entry.configure(state='disabled')
            self.signal_strong_entry.configure(state='disabled')
            self.signal_good_entry.configure(state='disabled')
            self.signal_normal_entry.configure(state='disabled')
            self.unlock_pwd_entry.configure(state='disabled')
            self.unlock_btn.configure(text='解锁参数')

    def start_test(self):
        global report_time, signal_threshold_strong, signal_threshold_good, signal_threshold_normal, param_unlock_password
        port_name = self.port_var.get().strip()
        if not port_name:
            messagebox.showwarning('提示', '请先选择串口')
            return

        try:
            report_time = int(self.wait_var.get().strip())
            if report_time <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning('提示', '上报等待时间必须是正整数')
            return

        try:
            strong = int(self.signal_strong_var.get().strip())
            good = int(self.signal_good_var.get().strip())
            normal = int(self.signal_normal_var.get().strip())
        except ValueError:
            messagebox.showwarning('提示', '信号阶梯值必须是整数')
            return

        if not (strong > good > normal):
            messagebox.showwarning('提示', '信号阶梯需满足: 阶梯1 > 阶梯2 > 阶梯3')
            return

        new_pwd = self.unlock_pwd_var.get().strip()
        if not new_pwd:
            messagebox.showwarning('提示', '解锁密码不能为空')
            return

        signal_threshold_strong = strong
        signal_threshold_good = good
        signal_threshold_normal = normal
        param_unlock_password = new_pwd

        try:
            open_serial_port(port_name, self.baud_var.get().strip() or '9600')
        except Exception as ex:
            messagebox.showerror('串口打开失败', str(ex))
            return

        stop_event.clear()
        self.is_running = True
        self.start_btn.configure(state='disabled')
        self.stop_btn.configure(state='normal')
        self._clear_table_data()
        self._reset_summary_panel()
        self.countdown_var.set(f'等待倒计时: {report_time}s')
        self.wait_phase_var.set('等待进度: 准备开始')
        self.wait_progress.configure(maximum=max(report_time, 1))
        self.wait_progress['value'] = 0
        self.status_var.set(f'测试中: {port_name}, 等待时间 {report_time}s')
        threading.Thread(target=self._run_test_in_thread, daemon=True).start()

    def _run_test_in_thread(self):
        try:
            print('\n' + '=' * 30 + ' 开始测试 ' + '=' * 30)
            enable_valve_test = bool(self.enable_valve_test_var.get())
            run_report_flow(
                on_before_snapshot=lambda data: self.root.after(0, self._update_before_snapshot, data),
                on_after_snapshot=lambda data: self.root.after(0, self._update_after_snapshot, data),
                on_summary_ready=lambda info: self.root.after(0, self._update_summary_panel, info),
                on_wait_progress=lambda phase, remaining, total, retry: self.root.after(
                    0, self._update_wait_progress, phase, remaining, total, retry
                ),
                enable_valve_test=enable_valve_test
            )
            print('=' * 30 + ' 测试完成 ' + '=' * 30 + '\n')
            self.root.after(0, self.status_var.set, '测试完成')
        except Exception as ex:
            print(f'测试异常: {ex}')
            err_info = dict(summary_info)
            if '失败' in str(ex) or '无通讯' in str(ex):
                err_info['report_result'] = '上报失败'
            else:
                err_info['report_result'] = '测试异常'
            err_info['final_message'] = str(ex)
            self.root.after(0, self._update_summary_panel, err_info)
            self.root.after(0, self.countdown_var.set, '等待倒计时: --')
            self.root.after(0, self.wait_phase_var.set, '等待进度: 异常中断')
            self.root.after(0, lambda: self.wait_progress.configure(maximum=100, value=0))
            self.root.after(0, self.status_var.set, '测试异常，请查看日志')
        finally:
            self.is_running = False
            close_serial_port()
            self.root.after(200, self.auto_save_log)
            self.root.after(260, self.auto_save_excel_result)
            self.root.after(0, lambda: self.start_btn.configure(state='normal'))
            self.root.after(0, lambda: self.stop_btn.configure(state='disabled'))

    def force_stop(self):
        if not self.is_running:
            return
        stop_event.set()
        self.status_var.set('正在强制停止...')
        self.wait_phase_var.set('等待进度: 用户强制停止')

    def clear_log(self):
        self.log_text.configure(state='normal')
        self.log_text.delete('1.0', tk.END)
        self.log_text.configure(state='disabled')

    def save_log(self):
        log_content = self.log_text.get('1.0', tk.END).strip()
        if not log_content:
            messagebox.showinfo('提示', '当前没有可保存的日志')
            return

        default_name = datetime.now().strftime('report_log_%Y%m%d_%H%M%S.txt')
        file_path = filedialog.asksaveasfilename(
            title='保存日志',
            defaultextension='.txt',
            initialfile=default_name,
            filetypes=[('文本文件', '*.txt'), ('所有文件', '*.*')]
        )
        if not file_path:
            return

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(log_content + '\n')
        self.status_var.set(f'日志已保存: {file_path}')

    def auto_save_log(self):
        log_content = self.log_text.get('1.0', tk.END).strip()
        if not log_content:
            return

        target_dir = 'D:\\Test_Result'
        os.makedirs(target_dir, exist_ok=True)

        meter_name = current_meter_no if current_meter_no else 'UNKNOWN'
        safe_meter_name = ''.join(c if c.isalnum() or c in ('-', '_') else '_' for c in meter_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_path = os.path.join(target_dir, f'{safe_meter_name}_{timestamp}.txt')

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(log_content + '\n')

        # 同一表号仅保留最新日志文件。
        prefix = f'{safe_meter_name}_'
        for name in os.listdir(target_dir):
            if not (name.startswith(prefix) and name.endswith('.txt')):
                continue
            old_path = os.path.join(target_dir, name)
            if os.path.normcase(old_path) == os.path.normcase(file_path):
                continue
            try:
                os.remove(old_path)
            except OSError:
                pass

        print(f'日志已自动保存: {file_path}')
        self.status_var.set(f'日志已自动保存: {file_path}')

    def auto_save_excel_result(self):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            print('未安装openpyxl，无法写入Excel结果')
            self.status_var.set('未安装openpyxl，无法写入Excel结果')
            return

        target_dir = 'D:\\Test_Result'
        os.makedirs(target_dir, exist_ok=True)

        # 按需求：每天新建一个Excel，文件名为 日期+Report_Result
        excel_name = datetime.now().strftime('%Y%m%d') + '_Report_Result.xlsx'
        excel_path = os.path.join(target_dir, excel_name)

        headers = ['时间', '表号', 'ABNT', '信号强度', '信号等级', '信噪比', 'IMEI', 'CCID', '成功次数', '上报结果', '关阀结果', '开阀结果']
        row_data = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            current_meter_no,
            summary_info.get('abnt', ''),
            summary_info.get('rsrp', ''),
            summary_info.get('signal_level', ''),
            summary_info.get('snr', ''),
            summary_info.get('imei', ''),
            summary_info.get('ccid', ''),
            summary_info.get('success_count', ''),
            summary_info.get('report_result', ''),
            summary_info.get('close_valve_result', '未执行'),
            summary_info.get('open_valve_result', '未执行')
        ]

        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            # 兼容当天已生成旧版本表头的场景，补齐新增列。
            for idx, header in enumerate(headers, start=1):
                if ws.cell(row=1, column=idx).value != header:
                    ws.cell(row=1, column=idx, value=header)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Result'
            ws.append(headers)

        # 同表号也保留历史：每次测试都追加一行，不覆盖旧记录。
        ws.append(row_data)
        status_text = f'结果已保存到Excel: {excel_path}'

        # 自动调整列宽，尽量保证一眼能看到完整内容。
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
            max_len = 0
            for cell in col_cells:
                if cell.value is None:
                    continue
                value_len = len(str(cell.value))
                if value_len > max_len:
                    max_len = value_len

            # 对中文和长串字段做适度放大，并设置上下限避免过窄或过宽。
            width = min(max(12, int(max_len * 1.6) + 2), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(excel_path)
        print(status_text)
        self.status_var.set(status_text)

    def on_close(self):
        if self.is_running:
            confirm = messagebox.askyesno('确认退出', '当前测试正在运行，确认强制停止并退出吗？')
        else:
            confirm = messagebox.askyesno('确认退出', '确认关闭上位机吗？')

        if not confirm:
            return

        stop_event.set()
        close_serial_port()
        sys.stdout = self.original_stdout
        sys.stderr = self.original_stderr
        self.root.destroy()


def main():
    root = tk.Tk()
    UpperComputerApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()



